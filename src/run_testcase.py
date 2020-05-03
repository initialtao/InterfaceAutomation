import os
import openpyxl
from interface_test import InterfaceTest

class RunTestcase:


#获取当前table路径
    def pathreplace(self,filename):

        path1 = os.path.abspath('..')
        print("path1：", path1)

        path2 = path1.replace('\\', '/')  # 替换反斜杠\要改成\\才能识别
        print("path2：", path2)
        # path='C:/Users/test/tao/InterfaceAutomation/testcase/testcase.xls'
        path2 = str(path2)
        path3 = os.path.join(path2, filename)
        print("path3是：", path3)
        path = path3.replace('\\', '/')
        print("path是：", path)
        return path

#执行table中的用例
    def run(self,path):
        #test_table = openpyxl.load_workbook('C:/Users/test/tao/InterfaceAutomation/testcase/testcase.xlsx')
        path = str(path)
        print(path)
        test_table = openpyxl.load_workbook(path)
        print("表格有：",test_table.sheetnames,test_table.active)
        test_table_sheet_title = test_table.sheetnames[0]
        print(test_table_sheet_title)
        testcase_table = test_table[test_table_sheet_title]
        print(testcase_table.title)
        #print(testcase_table.max_row)
        for i in range(2,testcase_table.max_row+1):
            print(i-1)
            testMethods = testcase_table.cell(i,2).value
            #print(testMethods)
            testUrl = str(testcase_table.cell(i,3).value)
            #print(testUrl)
            # params_type = testcase_table.cell(i, 3).value
            # print(params_type)
            testData = eval(testcase_table.cell(i,4).value)
            #print(type(testData))
            testCookies = str(testcase_table.cell(i, 5).value)
            #print(testCookies)
            tester = InterfaceTest()
            testcase_table.cell(i,6).value=tester.interface_test(testMethods,testUrl,testData,testCookies)
            print(testcase_table.cell(i,6).value)
        test_table.save(path)







if __name__ == '__main__':
    tt = RunTestcase()
    filename ='testcase/testcase.xlsx'
    path = tt.pathreplace(filename)
    #path ='C:\\Users\\test\tao\InterfaceAutomation\testcase\testcase.xlsx'
    tt.run(path)




